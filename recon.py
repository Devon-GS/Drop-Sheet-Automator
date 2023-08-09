import os
import pandas as pd
from openpyxl import load_workbook
from tkinter import filedialog

# Specify the folder path where the Excel files are located
folder_path = filedialog.askdirectory()
# folder_path = r"D:\1 - Coding and Course Work\My Apps\5 - Projects\Dad - Drop Recon\Excel Sheets"

# Collect all Excel files in the folder
excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]

# Check for ecel temp files
excel_files_filtered = []
for x in excel_files:
    # print(x)
    if not x.startswith('~$'):
        excel_files_filtered.append(x)
        
# Specify the output file path
output_file = filedialog.askopenfilename()
# output_file = r"D:\1 - Coding and Course Work\My Apps\5 - Projects\Dad - Drop Recon\Drops\06 CASH DROPS DAILY August 2023.xlsx"

# Load the output workbook
output_workbook = load_workbook(output_file)

# Starting row for writing data
output_worksheet = output_workbook.active
start_row = 5
start_column = 5  # Column E

# Loop through each Excel file
for file in excel_files_filtered:
    file_path = os.path.join(folder_path, file)

    # Read the Excel file into a pandas DataFrame
    df = pd.read_excel(file_path)

    # Retrieve the values of cell C2 and D2
    cell_b12_value = df.iloc[10, 1]
    cell_d12_value = df.iloc[10, 3]
    cell_d46_value = df.iloc[44, 3]

     # Write the values to the output worksheet
    output_worksheet.cell(row=start_row, column=start_column, value=cell_b12_value)
    output_worksheet.cell(row=start_row, column=start_column + 1, value=0)
    output_worksheet.cell(row=start_row, column=start_column + 2, value=cell_d12_value)
    output_worksheet.cell(row=start_row, column=start_column + 3, value=0)
    output_worksheet.cell(row=start_row, column=start_column + 5, value=cell_d46_value)

    # Increment the row for the next iteration
    start_row += 1

    # Print the values
    print(f"File: {file}")
    print(f"Cashier 1 Drop: {cell_b12_value}")
    print(f"Cashier 2 Drop: {cell_d12_value}")
    print(f"Management Drop: {cell_d46_value}")
    print()

    # Save the changes and close the output workbook
    output_workbook.save(output_file)
    output_workbook.close()

# Add a user input prompt to keep the console window open
print()
print("Finished processing")
input("Press Enter to exit...")
